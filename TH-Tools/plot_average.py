# Copyright (c) 2024, John Simonis and The Ohio State University
# This code was written by John Simonis for the ThunderHead research project at The Ohio State University.

# Python modules required by the current program
import os
import argparse
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

# A function to read in data from an excel spreadsheet.
def read_sheet_data(workbook, sheet_name):
    """Read the sheet data into a 2D list of floats."""
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        if all(isinstance(cell, (int, float)) for cell in row):
            data.append(list(row))
    return np.array(data)

# A function to calculate how close a matrix is to being all zeros, or in this case all candles extinguished.
def calculate_zero_proximity_score(matrix):
    """Calculate a percentage score indicating how close the matrix is to all zeros."""
    zero_proximity = 100 - (np.abs(matrix).mean() * 100)
    return max(0, zero_proximity)

# Function to average the excel matrices and then plot them using sns and matplotlib.
def plot_average_heatmaps(excel_files, output_dir):
    # Load the workbooks
    workbooks = [openpyxl.load_workbook(file, data_only=True) for file in excel_files]

    # Ensure all workbooks have the same sheets
    sheet_names = workbooks[0].sheetnames
    for wb in workbooks:
        if wb.sheetnames != sheet_names:
            raise ValueError("All Excel files must have the same sheet names in the same order.")

    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for sheet_name in sheet_names:
        # Read and average the sheet data from all workbooks
        matrices = [read_sheet_data(wb, sheet_name) for wb in workbooks]
        avg_matrix = np.mean(matrices, axis=0)

        # Handle all-zero matrices: Set the color scale to have a minimum below zero
        plt.figure(figsize=(10, 8))
        sns.heatmap(avg_matrix, cmap="YlGnBu", annot=True, fmt=".1f", cbar=True, vmin=0, vmax=1)

        # Calculate the zero proximity score
        zero_proximity_score = calculate_zero_proximity_score(avg_matrix)

        # Display the zero proximity score on the plot
        plt.title(f"Averaged Heatmap of {sheet_name}\nZero Proximity Score: {zero_proximity_score:.2f}%")

        # Save the plot
        output_file = os.path.join(output_dir, f"{sheet_name}_average_heatmap.png")
        plt.savefig(output_file)
        plt.close()
        print(f"Averaged heatmap saved for {sheet_name} at {output_file}")

if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Plot averaged heatmaps for corresponding sheets in multiple Excel files.')
    parser.add_argument('excel_files', type=str, nargs=3, help='Paths to the three Excel files.')
    parser.add_argument('output_dir', type=str, help='Directory to save the averaged heatmaps.')

    # Parse the arguments
    args = parser.parse_args()

    # Plot averaged heatmaps
    plot_average_heatmaps(args.excel_files, args.output_dir)

