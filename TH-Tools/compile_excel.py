# Copyright (c) 2024, John Simonis and The Ohio State University
# This code was written by John Simonis for the ThunderHead research project at The Ohio State University.

# Python modules required by the current program
import os
import re
from openpyxl import load_workbook, Workbook
import argparse

# Simple regex to remove various project titles.
def extract_numbers(string):
    """Extract all numbers from the string, returning a tuple of floats for sorting."""
    numbers = re.findall(r'\d+', string)
    return tuple(float(num) for num in numbers) if numbers else (float('inf'),)

# Function to compile all the individual matrices into one excel file.
def compile_excel_files(folder_path, output_file):
    # Create a new workbook
    compiled_workbook = Workbook()
    compiled_workbook.remove(compiled_workbook.active)  # Remove the default sheet

    # Get all Excel files in the folder
    excel_files = sorted([f for f in os.listdir(folder_path) if f.endswith((".xlsx", ".xls"))])

    tabs = []

    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        # Load the Excel file
        workbook = load_workbook(file_path)

        # Extract the part of the file name after the hyphen
        stripped_name = os.path.splitext(file_name)[0].split('-')[-1].strip()

        # Add each sheet and associated info to the tabs list
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            tabs.append((stripped_name, sheet))

    # Sort tabs by the numeric values extracted from the stripped name
    tabs.sort(key=lambda x: extract_numbers(x[0]))

    # Create new sorted sheets in the compiled workbook and copy content
    for name, original_sheet in tabs:
        new_sheet = compiled_workbook.create_sheet(title=name)
        for row in original_sheet.iter_rows(values_only=True):
            new_sheet.append(row)

    # Save the compiled workbook
    compiled_workbook.save(output_file)
    print(f"All Excel files compiled into {output_file}")

if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Compile multiple Excel files into one with separate sheets.')
    parser.add_argument('folder_path', type=str, help='Path to the folder containing Excel files.')
    parser.add_argument('output_file', type=str, help='Path to save the compiled Excel file.')

    # Parse the arguments
    args = parser.parse_args()

    # Compile the Excel files
    compile_excel_files(args.folder_path, args.output_file)

