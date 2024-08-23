# Copyright (c) 2024, John Simonis and The Ohio State University
# This code was written by John Simonis for the ThunderHead research project at The Ohio State University.

# Python modules required by the current program
import os
import argparse
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

# A function to read in data from an excel spreadsheet.
def read_sheet_data(sheet):
    """Read the sheet data into a 2D list of floats."""
    data = []
    for row in sheet.iter_rows(values_only=True):
        if all(isinstance(cell, (int, float)) for cell in row):
            data.append(list(row))
    return np.array(data)

# A function to calculate how close a matrix is to being all zeros, or in this case all candles extinguished.
def calculate_zero_proximity_score(matrix):
    """Calculate a percentage score indicating how close the matrix is to all zeros."""
    zero_proximity = 100 - (np.abs(matrix).mean() * 100)
    return max(0, zero_proximity)

# Function to plot the excel matrix using sns and matplotlib.
def plot_heatmaps(excel_file, output_dir):
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Read sheet data into a 2D list (assuming the sheet is a grid of numbers)
        data = read_sheet_data(sheet)

        if data.size > 0:
            # Handle all-zero matrices: Set the color scale to have a minimum below zero
            plt.figure(figsize=(10, 8))
            sns.heatmap(data, cmap="YlGnBu", annot=True, fmt=".1f", cbar=True, vmin=0, vmax=1)

            # Calculate the zero proximity score
            zero_proximity_score = calculate_zero_proximity_score(data)

            # Display the zero proximity score on the plot
            plt.title(f"Heatmap of {sheet_name}\nZero Proximity Score: {zero_proximity_score:.2f}%")

            # Save the plot
            output_file = os.path.join(output_dir, f"{sheet_name}_heatmap.png")
            plt.savefig(output_file)
            plt.close()
            print(f"Heatmap saved for {sheet_name} at {output_file}")

if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Plot heatmaps for each sheet in an Excel file with zero proximity scores.')
    parser.add_argument('excel_file', type=str, help='Path to the Excel file.')
    parser.add_argument('output_dir', type=str, help='Directory to save the heatmaps.')

    # Parse the arguments
    args = parser.parse_args()

    # Plot heatmaps
    plot_heatmaps(args.excel_file, args.output_dir)

